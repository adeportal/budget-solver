"""
Market Intelligence Excel sheet builder.

Consolidates all contextual signals applied to revenue forecasts into one sheet:
  1. Forecast Adjustments — holiday and weather correction factors per account
  2. Calibration Quality  — confidence score and active-day ratio per account
  3. Competitive Landscape — auction insights (trailing vs prior IS per competitor)
  4. Simulator Cross-Check — model prediction vs Google bid simulator at recommended spend

Each section uses traffic-light colours (green / amber / red) so a stakeholder
can scan the sheet in under 30 seconds and know which accounts need manual review.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from budget_solver.constants import NAV, GRN, RED, LBLU, WHIT
from budget_solver.excel.styling import _border

# Local colour constants
_AMBR = 'FFC000'
_LGRY = 'F2F2F2'
_DGRY = '808080'

_TITLE_FONT  = Font(name='Calibri', size=14, bold=True, color=NAV)
_SECTION_FONT = Font(name='Calibri', size=10, bold=True, color=WHIT)
_SECTION_FILL = PatternFill('solid', fgColor=NAV)
_HDR_FONT    = Font(name='Calibri', size=9, bold=True, color=WHIT)
_HDR_FILL    = PatternFill('solid', fgColor='2E75B6')
_DATA_FONT   = Font(name='Calibri', size=9)
_NOTE_FONT   = Font(name='Calibri', size=8, italic=True, color=_DGRY)
_ALT_FILL    = PatternFill('solid', fgColor=LBLU)
_GREEN_FILL  = PatternFill('solid', fgColor=GRN)
_AMBER_FILL  = PatternFill('solid', fgColor=_AMBR)
_RED_FILL    = PatternFill('solid', fgColor=RED)


def _cell(ws, row, col, value, font=None, fill=None, number_format=None,
          alignment=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:        c.font = font
    if fill:        c.fill = fill
    if number_format: c.number_format = number_format
    if alignment:   c.alignment = alignment
    if border:      c.border = border
    return c


def _section_header(ws, row, col, title, n_cols):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + n_cols - 1)
    c = ws.cell(row=row, column=col, value=title)
    c.font  = _SECTION_FONT
    c.fill  = _SECTION_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    return row + 1


def _traffic_fill(value: float, warn_lo: float, warn_hi: float,
                  good_direction: str = 'mid') -> PatternFill:
    """
    Return green/amber/red fill.
    good_direction:
      'mid'  — green when value is near 1.0, amber outside ±warn, red outside ±warn_hi
      'high' — green when high, red when low
      'low'  — green when low, red when high
    """
    if good_direction == 'mid':
        deviation = abs(value - 1.0)
        if deviation < warn_lo:   return _GREEN_FILL
        if deviation < warn_hi:   return _AMBER_FILL
        return _RED_FILL
    elif good_direction == 'high':
        if value >= warn_hi:  return _GREEN_FILL
        if value >= warn_lo:  return _AMBER_FILL
        return _RED_FILL
    else:  # low
        if value <= warn_lo:  return _GREEN_FILL
        if value <= warn_hi:  return _AMBER_FILL
        return _RED_FILL


def _build_market_intelligence(
    wb,
    forecast_label: str,
    accounts: list[str],
    holiday_corrections: dict,    # {acc: (factor, explanation)}
    weather_corrections: dict,    # {acc: (factor, explanation)}
    cal_confidence: dict,         # {acc: confidence_float}
    calibration_factors: dict,    # {acc: scale_float}
    auction_insights_df: pd.DataFrame,
    simulator_df: pd.DataFrame,
    predict_fns: dict,
    current_alloc: dict,
    recommended_alloc: dict,
):
    """Build the Market Intelligence sheet."""
    from budget_solver.bid_simulator import build_account_simulator_curve

    ws = wb.create_sheet('Market Intelligence')
    ws.sheet_view.showGridLines = False

    # Column widths
    col_widths = [32, 10, 45, 10, 45, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Sheet title ──────────────────────────────────────────
    ws.merge_cells(f'A{row}:F{row}')
    _cell(ws, row, 1, f'MARKET INTELLIGENCE — {forecast_label}',
          font=_TITLE_FONT,
          alignment=Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(f'A{row}:F{row}')
    _cell(ws, row, 1,
          'Contextual signals applied to revenue forecasts. '
          'Green = no concern | Amber = monitor | Red = review before acting.',
          font=_NOTE_FONT,
          alignment=Alignment(horizontal='left', vertical='center'))
    row += 2

    # ════════════════════════════════════════════════════════
    # SECTION 1: FORECAST ADJUSTMENTS
    # ════════════════════════════════════════════════════════
    row = _section_header(ws, row, 1, '1 — FORECAST ADJUSTMENTS', 6)

    hdrs = ['Account', 'Holiday ×', 'Holiday reason',
            'Weather ×', 'Weather reason', 'Combined ×']
    for c, h in enumerate(hdrs, 1):
        _cell(ws, row, c, h, font=_HDR_FONT, fill=_HDR_FILL,
              alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
              border=_border())
    row += 1

    for i, acc in enumerate(sorted(accounts)):
        hc, h_expl = holiday_corrections.get(acc, (1.0, '—'))
        wc, w_expl = weather_corrections.get(acc, (1.0, '—'))
        combined   = hc * wc
        fill = _ALT_FILL if i % 2 == 0 else PatternFill('solid', fgColor=WHIT)

        _cell(ws, row, 1, acc,      font=_DATA_FONT, fill=fill, border=_border())
        c2 = _cell(ws, row, 2, round(hc, 3), font=_DATA_FONT, border=_border(),
                   number_format='0.00"×"',
                   alignment=Alignment(horizontal='center'))
        c2.fill = _traffic_fill(hc, 0.05, 0.15, 'mid')
        _cell(ws, row, 3, h_expl,   font=_DATA_FONT, fill=fill, border=_border(),
              alignment=Alignment(wrap_text=True))
        c4 = _cell(ws, row, 4, round(wc, 3), font=_DATA_FONT, border=_border(),
                   number_format='0.00"×"',
                   alignment=Alignment(horizontal='center'))
        c4.fill = _traffic_fill(wc, 0.03, 0.10, 'mid')
        _cell(ws, row, 5, w_expl,   font=_DATA_FONT, fill=fill, border=_border(),
              alignment=Alignment(wrap_text=True))
        c6 = _cell(ws, row, 6, round(combined, 3), font=Font(name='Calibri', size=9, bold=True),
                   border=_border(), number_format='0.00"×"',
                   alignment=Alignment(horizontal='center'))
        c6.fill = _traffic_fill(combined, 0.06, 0.18, 'mid')
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # ════════════════════════════════════════════════════════
    # SECTION 2: CALIBRATION QUALITY
    # ════════════════════════════════════════════════════════
    row = _section_header(ws, row, 1, '2 — CALIBRATION QUALITY', 6)

    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 36

    hdrs2 = ['Account', 'Cal. factor', 'Confidence', 'Status', '', '']
    for c, h in enumerate(hdrs2, 1):
        _cell(ws, row, c, h if h else None,
              font=_HDR_FONT, fill=_HDR_FILL,
              alignment=Alignment(horizontal='center', vertical='center'),
              border=_border() if h else None)
    row += 1

    for i, acc in enumerate(sorted(accounts)):
        conf   = cal_confidence.get(acc, 1.0)
        factor = calibration_factors.get(acc, 1.0)
        fill   = _ALT_FILL if i % 2 == 0 else PatternFill('solid', fgColor=WHIT)

        if conf >= 0.7:
            status     = 'Clean window — full calibration applied'
            conf_fill  = _GREEN_FILL
        elif conf >= 0.4:
            status     = 'Moderate confidence — calibration partially applied'
            conf_fill  = _AMBER_FILL
        else:
            status     = 'Low confidence — paused or volatile window; model curve used'
            conf_fill  = _RED_FILL

        _cell(ws, row, 1, acc,          font=_DATA_FONT, fill=fill, border=_border())
        c2 = _cell(ws, row, 2, round(factor, 3),
                   font=_DATA_FONT, border=_border(),
                   number_format='0.000"×"',
                   alignment=Alignment(horizontal='center'))
        c2.fill = _traffic_fill(factor, 0.05, 0.15, 'mid')
        c3 = _cell(ws, row, 3, conf,
                   font=_DATA_FONT, border=_border(),
                   number_format='0%',
                   alignment=Alignment(horizontal='center'))
        c3.fill = conf_fill
        ws.merge_cells(start_row=row, start_column=4,
                       end_row=row,   end_column=6)
        _cell(ws, row, 4, status,       font=_DATA_FONT, fill=fill, border=_border(),
              alignment=Alignment(wrap_text=True))
        row += 1

    row += 1

    # ════════════════════════════════════════════════════════
    # SECTION 3: COMPETITIVE LANDSCAPE
    # ════════════════════════════════════════════════════════
    row = _section_header(ws, row, 1, '3 — COMPETITIVE LANDSCAPE (auction insights)', 6)

    # Reset column widths for this section
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 24

    if auction_insights_df is None or auction_insights_df.empty:
        ws.merge_cells(f'A{row}:F{row}')
        _cell(ws, row, 1,
              'No auction insights data. Run budget-solver-pull to fetch competitor IS data.',
              font=_NOTE_FONT)
        row += 2
    else:
        hdrs3 = ['Account', 'Competitor', 'Trailing IS', 'Prior IS', 'Δ IS', 'Signal']
        for c, h in enumerate(hdrs3, 1):
            _cell(ws, row, c, h, font=_HDR_FONT, fill=_HDR_FILL,
                  alignment=Alignment(horizontal='center', vertical='center'),
                  border=_border())
        row += 1

        for acc in sorted(auction_insights_df['account_name'].unique()):
            acc_rows = (auction_insights_df[auction_insights_df['account_name'] == acc]
                        .sort_values('trailing_is', ascending=False))
            for j, r in enumerate(acc_rows.itertuples()):
                fill = _ALT_FILL if j % 2 == 0 else PatternFill('solid', fgColor=WHIT)
                surge = r.is_delta > 0.10
                drop  = r.is_delta < -0.10
                signal = '⚠ Competitor surging — expect CPC pressure' if surge else \
                         ('↓ Competitor retreating — opportunity' if drop else 'Stable')
                sig_fill = _RED_FILL if surge else (_GREEN_FILL if drop else fill)

                _cell(ws, row, 1, acc if j == 0 else '', font=_DATA_FONT, fill=fill, border=_border())
                _cell(ws, row, 2, r.domain, font=_DATA_FONT, fill=fill, border=_border())
                _cell(ws, row, 3, r.trailing_is, font=_DATA_FONT, fill=fill, border=_border(),
                      number_format='0%', alignment=Alignment(horizontal='center'))
                _cell(ws, row, 4, r.prior_is, font=_DATA_FONT, fill=fill, border=_border(),
                      number_format='0%', alignment=Alignment(horizontal='center'))
                d_cell = _cell(ws, row, 5, r.is_delta, font=_DATA_FONT, border=_border(),
                               number_format='+0%;-0%;0%',
                               alignment=Alignment(horizontal='center'))
                d_cell.fill = _traffic_fill(1.0 + r.is_delta, 0.05, 0.10, 'mid')
                _cell(ws, row, 6, signal, font=_DATA_FONT, fill=sig_fill, border=_border(),
                      alignment=Alignment(wrap_text=True))
                row += 1
            row += 0  # no gap between accounts; section gap after all

        row += 1

    # ════════════════════════════════════════════════════════
    # SECTION 4: SIMULATOR CROSS-CHECK
    # ════════════════════════════════════════════════════════
    row = _section_header(ws, row, 1, '4 — SIMULATOR CROSS-CHECK (model vs Google bid simulator)', 6)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 36

    if simulator_df is None or simulator_df.empty:
        ws.merge_cells(f'A{row}:F{row}')
        _cell(ws, row, 1,
              'No simulator data. Run budget-solver-pull to fetch bid simulation points.',
              font=_NOTE_FONT)
        row += 2
    else:
        hdrs4 = ['Account', 'Rec. spend', 'Model revenue', 'Simulator revenue', 'Δ', 'Signal']
        for c, h in enumerate(hdrs4, 1):
            _cell(ws, row, c, h, font=_HDR_FONT, fill=_HDR_FILL,
                  alignment=Alignment(horizontal='center', vertical='center'),
                  border=_border())
        row += 1

        for i, acc in enumerate(sorted(predict_fns.keys())):
            sim_fn = build_account_simulator_curve(
                simulator_df, acc, current_alloc.get(acc, 0.0)
            )
            if sim_fn is None:
                continue

            rec_sp    = recommended_alloc.get(acc, current_alloc.get(acc, 0.0))
            model_rev = predict_fns[acc](rec_sp)
            sim_rev   = sim_fn(rec_sp)

            if model_rev <= 0 or sim_rev <= 0:
                continue

            delta = (model_rev - sim_rev) / sim_rev
            fill  = _ALT_FILL if i % 2 == 0 else PatternFill('solid', fgColor=WHIT)

            if abs(delta) < 0.10:
                signal    = 'Good agreement'
                sig_fill  = _GREEN_FILL
            elif abs(delta) < 0.20:
                signal    = 'Minor divergence — monitor'
                sig_fill  = _AMBER_FILL
            elif delta > 0:
                signal    = '⚠ Model more optimistic than simulator — consider caution'
                sig_fill  = _RED_FILL
            else:
                signal    = '⚠ Model more conservative than simulator — may be under-forecasting'
                sig_fill  = _AMBER_FILL

            _cell(ws, row, 1, acc,       font=_DATA_FONT, fill=fill, border=_border())
            _cell(ws, row, 2, rec_sp,    font=_DATA_FONT, fill=fill, border=_border(),
                  number_format='€#,##0', alignment=Alignment(horizontal='right'))
            _cell(ws, row, 3, model_rev, font=_DATA_FONT, fill=fill, border=_border(),
                  number_format='€#,##0', alignment=Alignment(horizontal='right'))
            _cell(ws, row, 4, sim_rev,   font=_DATA_FONT, fill=fill, border=_border(),
                  number_format='€#,##0', alignment=Alignment(horizontal='right'))
            d_cell = _cell(ws, row, 5, delta, font=Font(name='Calibri', size=9, bold=True),
                           border=_border(), number_format='+0%;-0%;0%',
                           alignment=Alignment(horizontal='center'))
            d_cell.fill = _traffic_fill(1.0 + delta, 0.10, 0.20, 'mid')
            _cell(ws, row, 6, signal,    font=_DATA_FONT, fill=sig_fill, border=_border(),
                  alignment=Alignment(wrap_text=True))
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        _cell(ws, row, 1,
              'Simulator uses Google\'s internal auction model (forward-looking). '
              'Model uses historically-fitted response curves. '
              'Divergence > 20% warrants manual review before sharing with stakeholders.',
              font=_NOTE_FONT,
              alignment=Alignment(wrap_text=True))
        ws.row_dimensions[row].height = 28
