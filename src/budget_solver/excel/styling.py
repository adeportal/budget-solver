"""
Excel formatting and styling helpers.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from budget_solver.constants import REPORTING_ZERO_EPSILON, NAV, WHIT


def _hdr(ws, row, ncols, bg=NAV, fg=WHIT):
    """Apply header styling to a row."""
    fill = PatternFill('solid', fgColor=bg)
    font = Font(bold=True, color=fg, name='Calibri', size=10)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _fmt_col(ws, col, width):
    """Set column width."""
    ws.column_dimensions[get_column_letter(col)].width = width


def _border():
    """Return a thin border for cells."""
    thin = Side(border_style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def sanitize_display_currency(value, epsilon=REPORTING_ZERO_EPSILON):
    """Zero-out numerical noise that would otherwise display as €0."""
    value = float(value)
    return 0.0 if abs(value) < epsilon else value
