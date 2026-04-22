"""
Excel report generation for budget optimization scenarios.

Phase 6: Multi-scenario strategic presentation (A → B → C/C1 → D).
"""
from datetime import datetime

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference

from budget_solver.constants import NAV, BLUE, GRN, RED, LGRY, LBLU, WHIT
from budget_solver.excel.styling import _hdr, _fmt_col, _border, sanitize_display_currency
from budget_solver.narrative import full_scenario_narrative
from budget_solver.excel.builders import _build_overview, _build_scenario_sheet
from budget_solver.excel.phase6b import _build_extended_budget, _build_curve_diagnostics, _build_outlier_log, _build_demand_index


def build_excel(
    scenario_set,
    df: pd.DataFrame,
    output_path: str,
    model_info: dict = None,
    account_data: dict = None,
    removal_log: list = None,
    demand_index: dict = None,
    demand_normalized: bool = False,
    forecast_week: int = None,
    forecast_label: str = 'Forecast',
    actual_window_label: str = 'Last 30 days',
    actual_window_detail: str = None,
    extended_budget_steps: int = 6,
):
    """
    Build multi-scenario Excel report.

    Args:
        scenario_set: ScenarioSet with scenarios A/B/C/D
        df: Raw input data
        output_path: Path to save Excel file
        model_info: dict[account] → (fn, params, r2, mname)
        account_data: dict[account] → {'spend': [...], 'revenue': [...]}
        removal_log: List of outlier removal entries
        demand_index: dict[iso_week] → demand multiplier
        demand_normalized: Whether curves were demand-normalized
        forecast_week: ISO week for forecast period
        forecast_label: Label for forecast period (e.g., "May 2026")
        actual_window_label: Label for actual baseline window
        actual_window_detail: Detailed description of actual window
        extended_budget_steps: Number of rows in Extended Budget sweep

    Returns:
        output_path
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Extract scenarios
    scenarios = scenario_set.scenarios
    scen_a = scenarios[0]  # Current Run Rate
    scen_b = scenarios[1]  # Target Budget
    scen_c = scenarios[2]  # Recommended (C or C1)
    scen_d = scenarios[3] if len(scenarios) > 3 else None  # Max Justified

    min_mroas = scenario_set.min_mroas

    # Build sheets
    _build_executive_summary(wb, scenario_set, forecast_label, actual_window_label)
    _build_overview(wb, scenario_set)
    _build_scenario_sheet(wb, scen_a, None, min_mroas, "A")
    _build_scenario_sheet(wb, scen_b, scen_a, min_mroas, "B")
    _build_scenario_sheet(wb, scen_c, scen_b, min_mroas, "C")
    if scen_d:
        _build_scenario_sheet(wb, scen_d, scen_c, min_mroas, "D")
    _build_extended_budget(wb, scenario_set, extended_budget_steps)
    _build_curve_diagnostics(wb, scenario_set, model_info, account_data)
    _build_outlier_log(wb, removal_log if removal_log else [])
    _build_demand_index(wb, demand_index if demand_index else {}, demand_normalized, forecast_week if forecast_week else 1, forecast_label)

    wb.save(output_path)
    return output_path


def _build_executive_summary(wb, scenario_set, forecast_label, actual_window_label):
    """
    Build Executive Summary sheet — single-page brief for stakeholders.

    Layout:
    - Title block
    - Headline KPIs (3-column: Current A, Recommended C, Max D)
    - Methodology section
    - Scenario findings (4 narrative blocks from Phase 5)
    - Top 5 action recommendations
    """
    ws = wb.create_sheet('Executive Summary')
    ws.sheet_view.showGridLines = False

    scenarios = scenario_set.scenarios
    scen_a = scenarios[0]
    scen_b = scenarios[1]
    scen_c = scenarios[2]
    scen_d = scenarios[3] if len(scenarios) > 3 else None
    min_mroas = scenario_set.min_mroas

    row = 1

    # ── Title block ──
    ws.cell(row=row, column=1, value='BUDGET OPTIMIZATION — STRATEGIC RECOMMENDATIONS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, name='Calibri', color=NAV)
    row += 1

    ws.cell(row=row, column=1, value=(
        f'Generated: {scenario_set.generated_at.strftime("%Y-%m-%d %H:%M")}  |  '
        f'Minimum mROAS floor: {min_mroas:.1f}x  |  '
        f'Forecast period: {forecast_label}'
    ))
    ws.cell(row=row, column=1).font = Font(size=9, name='Calibri', color='888888')
    row += 2

    # ── Headline KPIs (5-column layout: label + 4 scenarios) ──
    # Column headers (start from column 2 to leave room for row labels)
    ws.cell(row=row, column=1, value='')  # Empty top-left cell
    kpi_headers = ['CURRENT (A)', 'TARGET (B)', f'RECOMMENDED (C)', 'MAX JUSTIFIED (D)']
    for c, hdr in enumerate(kpi_headers, 2):  # Start from column 2
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = Font(bold=True, size=10, name='Calibri', color=WHIT)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row += 1

    # KPI rows with row labels in column 1
    kpi_rows = [
        ('Budget',
         f'€{scen_a.budget_monthly:,.0f}',
         f'€{scen_b.budget_monthly:,.0f}',
         f'€{scen_c.budget_monthly:,.0f}',
         f'€{scen_d.budget_monthly:,.0f}' if scen_d else 'N/A'),
        ('Revenue',
         f'€{scen_a.revenue_monthly / 1_000_000:.2f}M',
         f'€{scen_b.revenue_monthly / 1_000_000:.2f}M',
         f'€{scen_c.revenue_monthly / 1_000_000:.2f}M',
         f'€{scen_d.revenue_monthly / 1_000_000:.2f}M' if scen_d else 'N/A'),
        ('Blended ROAS',
         f'{scen_a.blended_roas:.2f}x',
         f'{scen_b.blended_roas:.2f}x',
         f'{scen_c.blended_roas:.2f}x',
         f'{scen_d.blended_roas:.2f}x' if scen_d else 'N/A'),
        ('Uplift vs B',
         '—',
         '—',
         f'€{(scen_c.revenue_monthly - scen_b.revenue_monthly) / 1000:+,.0f}k ({(scen_c.revenue_monthly / scen_b.revenue_monthly - 1) * 100:+.1f}%)',
         f'€{(scen_d.revenue_monthly - scen_b.revenue_monthly) / 1000:+,.0f}k ({(scen_d.revenue_monthly / scen_b.revenue_monthly - 1) * 100:+.1f}%)' if scen_d else 'N/A'),
    ]

    for label, val_a, val_b, val_c, val_d in kpi_rows:
        # Row label in column 1
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=9, name='Calibri', color='444444')
        cell.alignment = Alignment(horizontal='right', vertical='center')

        # Values in columns 2, 3, 4, 5
        ws.cell(row=row, column=2, value=val_a).font = Font(size=10, name='Calibri', color=BLUE)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=3, value=val_b).font = Font(size=10, name='Calibri', color=BLUE)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=4, value=val_c).font = Font(size=10, name='Calibri', color=BLUE, bold=True)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=5, value=val_d).font = Font(size=10, name='Calibri', color=BLUE)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')

        row += 1

    row += 1

    # ── Methodology section ──
    ws.cell(row=row, column=1, value='APPROACH & METHODOLOGY')
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, name='Calibri', color=NAV)
    row += 1

    methodology = (
        f'Budget optimization uses fitted response curves (log/power models) to allocate spend across accounts, '
        f'maximizing total projected revenue for {forecast_label}. The optimizer (SLSQP) equalizes marginal ROI '
        f'across accounts, subject to a {min_mroas:.1f}x minimum instantaneous mROAS floor (breakeven constraint). '
        f'Curves are calibrated to actual {actual_window_label} lag-adjusted ROAS to anchor predictions to current reality.\n\n'
        f'Scenarios A→B→C→D represent a decision cascade: A = current run rate (baseline), '
        f'B = target budget allocated proportionally, C = recommended reallocation (optimized), '
        f'D = theoretical maximum (all accounts at {min_mroas:.1f}x floor). '
        f'Discrete mROAS between scenarios measures incremental return on marginal spend changes. '
        f'Phase 4 stability rules limit account changes to top-2 optional moves (mandatory floor caps always preserved).'
    )

    cell = ws.cell(row=row, column=1, value=methodology)
    cell.font = Font(size=9, name='Calibri', color='444444')
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 90
    row += 2

    # ── Scenario Findings ──
    ws.cell(row=row, column=1, value='SCENARIO FINDINGS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, name='Calibri', color=NAV)
    row += 1

    # Use Phase 5 narratives for each scenario
    prev_scenario = None
    for i, scen in enumerate(scenarios):
        # Scenario header
        recommended_mark = ' ✅ RECOMMENDED' if scen.recommended else ''
        ws.cell(row=row, column=1, value=f'SCENARIO {scen.id}: {scen.name.upper()}{recommended_mark}')
        ws.cell(row=row, column=1).font = Font(bold=True, size=10, name='Calibri', color=BLUE)
        row += 1

        # Get narrative (use full_scenario_narrative but extract just the summary and key points)
        narrative = full_scenario_narrative(scen, prev_scenario, min_mroas)

        # Extract just the summary paragraph and portfolio headline (skip per-account detail for exec summary)
        lines = narrative.split('\n')
        summary_lines = []
        in_summary = False
        for line in lines:
            if line.startswith('Budget:'):
                # Portfolio headline
                summary_lines.append(line)
                in_summary = True
            elif line.startswith('→'):
                # Summary paragraph
                summary_lines.append(line[2:].strip())  # Remove "→ " prefix
            elif in_summary and (line.startswith('Per-account:') or line.startswith('Action items:')):
                # Stop at per-account detail
                break
            elif in_summary and line.strip() and not line.startswith('⚠️'):
                # Additional summary lines (but skip warnings for exec summary brevity)
                if not line.startswith('  '):  # Skip indented warning detail
                    summary_lines.append(line.strip())

        summary_text = '\n'.join(summary_lines[:5])  # Limit to first 5 lines for brevity

        cell = ws.cell(row=row, column=1, value=summary_text)
        cell.font = Font(size=9, name='Calibri', color='444444')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 70  # Increased from 50 to prevent truncation
        row += 1

        prev_scenario = scen
        row += 1  # Gap between scenarios

    # ── Top 5 Action Recommendations ──
    ws.cell(row=row, column=1, value='TOP ACTION RECOMMENDATIONS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, name='Calibri', color=NAV)
    row += 1

    # Extract action items from Scenario C narrative
    c_narrative = full_scenario_narrative(scen_c, scen_b, min_mroas)
    action_lines = []
    in_actions = False
    for line in c_narrative.split('\n'):
        if line.startswith('Action items:'):
            in_actions = True
            continue
        if in_actions:
            if line.strip() and line.strip()[0].isdigit():
                # Action item line (starts with number)
                action_lines.append(line.strip())
            elif line.strip() and not line.strip().startswith('──'):
                # Continuation line (indented)
                if action_lines:
                    action_lines[-1] += ' ' + line.strip()

    # Take top 5 (or all if fewer)
    for i, action in enumerate(action_lines[:5], 1):
        # Remove the existing number prefix (e.g., "1. ") and re-number
        action_text = action.split('. ', 1)[1] if '. ' in action else action

        cell = ws.cell(row=row, column=1, value=f'{i}. {action_text}')
        cell.font = Font(size=9, name='Calibri', color='444444')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 28
        row += 1

    if not action_lines:
        cell = ws.cell(row=row, column=1, value='No significant account changes recommended (current allocation near-optimal).')
        cell.font = Font(size=9, name='Calibri', color='888888', italic=True)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 18  # Row labels
    ws.column_dimensions['B'].width = 24  # Scenario A
    ws.column_dimensions['C'].width = 24  # Scenario B
    ws.column_dimensions['D'].width = 28  # Scenario C (recommended, slightly wider)
    ws.column_dimensions['E'].width = 24  # Scenario D
