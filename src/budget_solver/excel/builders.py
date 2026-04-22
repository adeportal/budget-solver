"""
Helper functions for building individual Excel sheets (Phase 6).
"""
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

from budget_solver.constants import NAV, BLUE, GRN, RED, LBLU, WHIT
from budget_solver.excel.styling import _hdr, _fmt_col, _border
from budget_solver.narrative import full_scenario_narrative


def _build_overview(wb, scenario_set):
    """
    Build Overview sheet — side-by-side table comparing all scenarios.

    Layout:
    - Columns: [Metric | A | B | C | D]
    - Portfolio metrics: budget, revenue, ROAS, portfolio discrete mROAS
    - Per-account metrics: spend, revenue, discrete mROAS transitions, inst. mROAS
    - Conditional formatting on mROAS cells
    - Color legend at bottom
    """
    ws = wb.create_sheet('Overview')
    ws.sheet_view.showGridLines = False

    scenarios = scenario_set.scenarios
    scen_a = scenarios[0]
    scen_b = scenarios[1]
    scen_c = scenarios[2]
    scen_d = scenarios[3] if len(scenarios) > 3 else None
    min_mroas = scenario_set.min_mroas

    # Get unique accounts (sorted)
    accounts = sorted(set(a.account for s in scenarios for a in s.allocations))

    row = 1

    # ── Title ──
    ws.cell(row=row, column=1, value='SCENARIO OVERVIEW — SIDE-BY-SIDE COMPARISON')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, name='Calibri', color=NAV)
    row += 2

    # ── Header row ──
    headers = ['Metric', 'A — Current', 'B — Target', 'C — Recommended', 'D — Max @ Floor']
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = Font(bold=True, size=10, name='Calibri', color=WHIT)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row += 1
    start_data_row = row  # For conditional formatting later

    # ── Portfolio metrics ──
    portfolio_rows = [
        ('Monthly Budget', scen_a.budget_monthly, scen_b.budget_monthly, scen_c.budget_monthly, scen_d.budget_monthly if scen_d else None),
        ('Monthly Revenue', scen_a.revenue_monthly, scen_b.revenue_monthly, scen_c.revenue_monthly, scen_d.revenue_monthly if scen_d else None),
        ('Blended ROAS', scen_a.blended_roas, scen_b.blended_roas, scen_c.blended_roas, scen_d.blended_roas if scen_d else None),
        ('Portfolio Disc. mROAS vs Prev', None, scen_b.portfolio_discrete_mroas, scen_c.portfolio_discrete_mroas, scen_d.portfolio_discrete_mroas if scen_d else None),
    ]

    for label, val_a, val_b, val_c, val_d in portfolio_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9, name='Calibri', color='444444')

        # Format values
        if 'Budget' in label:
            fmt = '€#,##0'
        elif 'Revenue' in label:
            fmt = '€#,##0,,"M"'  # Millions for readability
        elif 'ROAS' in label or 'mROAS' in label:
            fmt = '0.00"x"'
        else:
            fmt = '@'

        for c, val in enumerate([val_a, val_b, val_c, val_d], 2):
            cell = ws.cell(row=row, column=c, value=val if val is not None else '—')
            if val is not None:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal='center')
            cell.border = _border()

        row += 1

    row += 1  # Gap

    # ── Per-account monthly spend ──
    cell = ws.cell(row=row, column=1, value='PER-ACCOUNT MONTHLY SPEND')
    cell.font = Font(bold=True, size=10, name='Calibri', color=WHIT)
    cell.fill = PatternFill('solid', fgColor=NAV)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    for account in accounts:
        ws.cell(row=row, column=1, value=account).font = Font(size=9, name='Calibri', color='444444')

        # Find allocation for this account in each scenario
        alloc_a = next((a for a in scen_a.allocations if a.account == account), None)
        alloc_b = next((a for a in scen_b.allocations if a.account == account), None)
        alloc_c = next((a for a in scen_c.allocations if a.account == account), None)
        alloc_d = next((a for a in scen_d.allocations if a.account == account), None) if scen_d else None

        for c, alloc in enumerate([alloc_a, alloc_b, alloc_c, alloc_d], 2):
            val = alloc.monthly_spend if alloc else None
            cell = ws.cell(row=row, column=c, value=val if val is not None else '—')
            if val is not None:
                cell.number_format = '€#,##0'
            cell.alignment = Alignment(horizontal='center')
            cell.border = _border()

        row += 1

    row += 1  # Gap

    # ── Per-account monthly revenue ──
    cell = ws.cell(row=row, column=1, value='PER-ACCOUNT MONTHLY REVENUE')
    cell.font = Font(bold=True, size=10, name='Calibri', color=WHIT)
    cell.fill = PatternFill('solid', fgColor=NAV)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    for account in accounts:
        ws.cell(row=row, column=1, value=account).font = Font(size=9, name='Calibri', color='444444')

        alloc_a = next((a for a in scen_a.allocations if a.account == account), None)
        alloc_b = next((a for a in scen_b.allocations if a.account == account), None)
        alloc_c = next((a for a in scen_c.allocations if a.account == account), None)
        alloc_d = next((a for a in scen_d.allocations if a.account == account), None) if scen_d else None

        for c, alloc in enumerate([alloc_a, alloc_b, alloc_c, alloc_d], 2):
            val = alloc.monthly_revenue if alloc else None
            cell = ws.cell(row=row, column=c, value=val if val is not None else '—')
            if val is not None:
                cell.number_format = '€#,##0'
            cell.alignment = Alignment(horizontal='center')
            cell.border = _border()

        row += 1

    row += 1  # Gap

    # ── Discrete mROAS transitions (separate row for each transition) ──
    cell = ws.cell(row=row, column=1, value='DISCRETE mROAS TRANSITIONS')
    cell.font = Font(bold=True, size=10, name='Calibri', color=WHIT)
    cell.fill = PatternFill('solid', fgColor=NAV)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    # A→B transition
    ws.cell(row=row, column=1, value='Disc. mROAS (A→B)').font = Font(size=9, name='Calibri', color='444444', italic=True)
    for account in accounts:
        alloc_b = next((a for a in scen_b.allocations if a.account == account), None)
        val = alloc_b.discrete_mroas_vs_prev if alloc_b else None
        # Only show in column B (column index 3)
        if val is not None and val != 0:
            cell = ws.cell(row=row, column=3, value=val)
            cell.number_format = '0.00"x"'
            cell.alignment = Alignment(horizontal='center')
        row += 1
    row -= len(accounts)  # Reset to start of this transition block

    # Fill in the account names in column 1 for each account row
    for i, account in enumerate(accounts):
        ws.cell(row=row + i + 1, column=1, value=f'  {account}').font = Font(size=9, name='Calibri', color='666666')

        alloc_b = next((a for a in scen_b.allocations if a.account == account), None)
        val_b = alloc_b.discrete_mroas_vs_prev if alloc_b else None

        # Column 2 (A): —
        ws.cell(row=row + i + 1, column=2, value='—').alignment = Alignment(horizontal='center')
        ws.cell(row=row + i + 1, column=2).border = _border()

        # Column 3 (B): disc mROAS A→B
        cell = ws.cell(row=row + i + 1, column=3, value=val_b if val_b is not None else '—')
        if val_b is not None:
            cell.number_format = '0.00"x"'
        cell.alignment = Alignment(horizontal='center')
        cell.border = _border()

        # Columns 4, 5 (C, D): —
        ws.cell(row=row + i + 1, column=4, value='—').alignment = Alignment(horizontal='center')
        ws.cell(row=row + i + 1, column=4).border = _border()
        ws.cell(row=row + i + 1, column=5, value='—').alignment = Alignment(horizontal='center')
        ws.cell(row=row + i + 1, column=5).border = _border()

    row += len(accounts) + 1

    # B→C transition
    ws.cell(row=row, column=1, value='Disc. mROAS (B→C)').font = Font(size=9, name='Calibri', color='444444', italic=True)
    row += 1

    for account in accounts:
        ws.cell(row=row, column=1, value=f'  {account}').font = Font(size=9, name='Calibri', color='666666')

        alloc_c = next((a for a in scen_c.allocations if a.account == account), None)
        val_c = alloc_c.discrete_mroas_vs_prev if alloc_c else None

        # Columns 2, 3 (A, B): —
        ws.cell(row=row, column=2, value='—').alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).border = _border()
        ws.cell(row=row, column=3, value='—').alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).border = _border()

        # Column 4 (C): disc mROAS B→C
        cell = ws.cell(row=row, column=4, value=val_c if val_c is not None else '—')
        if val_c is not None:
            cell.number_format = '0.00"x"'
        cell.alignment = Alignment(horizontal='center')
        cell.border = _border()

        # Column 5 (D): —
        ws.cell(row=row, column=5, value='—').alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5).border = _border()

        row += 1

    row += 1

    # C→D transition (if D exists)
    if scen_d:
        ws.cell(row=row, column=1, value='Disc. mROAS (C→D)').font = Font(size=9, name='Calibri', color='444444', italic=True)
        row += 1

        for account in accounts:
            ws.cell(row=row, column=1, value=f'  {account}').font = Font(size=9, name='Calibri', color='666666')

            alloc_d = next((a for a in scen_d.allocations if a.account == account), None)
            val_d = alloc_d.discrete_mroas_vs_prev if alloc_d else None

            # Columns 2, 3, 4 (A, B, C): —
            ws.cell(row=row, column=2, value='—').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2).border = _border()
            ws.cell(row=row, column=3, value='—').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).border = _border()
            ws.cell(row=row, column=4, value='—').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4).border = _border()

            # Column 5 (D): disc mROAS C→D
            cell = ws.cell(row=row, column=5, value=val_d if val_d is not None else '—')
            if val_d is not None:
                cell.number_format = '0.00"x"'
            cell.alignment = Alignment(horizontal='center')
            cell.border = _border()

            row += 1

        row += 1

    # ── Per-account inst. mROAS ──
    cell = ws.cell(row=row, column=1, value='PER-ACCOUNT INST. mROAS')
    cell.font = Font(bold=True, size=10, name='Calibri', color=WHIT)
    cell.fill = PatternFill('solid', fgColor=NAV)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    inst_mroas_start_row = row  # For conditional formatting

    for account in accounts:
        ws.cell(row=row, column=1, value=account).font = Font(size=9, name='Calibri', color='444444')

        alloc_a = next((a for a in scen_a.allocations if a.account == account), None)
        alloc_b = next((a for a in scen_b.allocations if a.account == account), None)
        alloc_c = next((a for a in scen_c.allocations if a.account == account), None)
        alloc_d = next((a for a in scen_d.allocations if a.account == account), None) if scen_d else None

        for c, alloc in enumerate([alloc_a, alloc_b, alloc_c, alloc_d], 2):
            val = alloc.inst_mroas if alloc else None
            cell = ws.cell(row=row, column=c, value=val if val is not None else '—')
            if val is not None:
                cell.number_format = '0.00"x"'
            cell.alignment = Alignment(horizontal='center')
            cell.border = _border()

        row += 1

    inst_mroas_end_row = row - 1

    # ── Apply conditional formatting to inst. mROAS and discrete mROAS cells ──
    # Inst. mROAS: columns B-E, rows inst_mroas_start_row to inst_mroas_end_row
    for col_letter in ['B', 'C', 'D', 'E']:
        # Red: < min_mroas
        ws.conditional_formatting.add(
            f'{col_letter}{inst_mroas_start_row}:{col_letter}{inst_mroas_end_row}',
            CellIsRule(operator='lessThan', formula=[str(min_mroas)], fill=PatternFill(start_color=RED, end_color=RED, fill_type='solid'))
        )
        # Yellow: >= min_mroas and < min_mroas + 1.5
        ws.conditional_formatting.add(
            f'{col_letter}{inst_mroas_start_row}:{col_letter}{inst_mroas_end_row}',
            CellIsRule(operator='between', formula=[str(min_mroas), str(min_mroas + 1.5)], fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'))
        )
        # Green: >= min_mroas + 1.5
        ws.conditional_formatting.add(
            f'{col_letter}{inst_mroas_start_row}:{col_letter}{inst_mroas_end_row}',
            CellIsRule(operator='greaterThanOrEqual', formula=[str(min_mroas + 1.5)], fill=PatternFill(start_color=GRN, end_color=GRN, fill_type='solid'))
        )

    # TODO: Apply conditional formatting to discrete mROAS cells (more complex due to scattered layout)

    row += 1

    # ── Color legend ──
    ws.cell(row=row, column=1, value='mROAS Color Legend:').font = Font(size=9, name='Calibri', color='444444', italic=True)
    row += 1

    legend_items = [
        (f'🔴 Red: < {min_mroas:.1f}x (below floor)', RED),
        (f'🟡 Yellow: {min_mroas:.1f}x - {min_mroas + 1.5:.1f}x (monitor)', 'FFC000'),
        (f'🟢 Green: ≥ {min_mroas + 1.5:.1f}x (healthy)', GRN),
    ]

    for label, color in legend_items:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(size=8, name='Calibri', color='666666')
        # cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22  # Recommended (C), slightly wider
    ws.column_dimensions['E'].width = 18


def _build_scenario_sheet(wb, scenario, prev_scenario, min_mroas, sheet_suffix):
    """
    Build a per-scenario sheet (A, B, C, D).

    Layout:
    - Narrative block from Phase 5 at top
    - Per-account table: Account | Daily Spend | Monthly Spend | Daily Revenue | Monthly Revenue | ROAS | Inst. mROAS | Disc. mROAS vs Prev | Move
    - Conditional formatting on mROAS columns
    """
    from budget_solver.constants import DAYS_PER_MONTH

    # Sheet name
    if scenario.id == "C1":
        sheet_name = "C1 - Budget-Neutral"
    else:
        sheet_name = f"{scenario.id} - {scenario.name}"

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Get narrative from Phase 5
    narrative = full_scenario_narrative(scenario, prev_scenario, min_mroas)

    # Write narrative as merged cells at top
    row = 1
    lines = narrative.split('\n')

    for line in lines:
        if not line.strip():
            row += 1
            continue

        # Merge across all columns for readability
        cell = ws.cell(row=row, column=1, value=line)

        # Styling based on line type
        if line.startswith('──'):
            # Header line
            cell.font = Font(bold=True, size=12, name='Calibri', color=NAV)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.row_dimensions[row].height = 18
        elif line.startswith('Budget:') or line.startswith('Revenue:') or line.startswith('ROAS:'):
            # Portfolio metrics
            cell.font = Font(size=9, name='Calibri', color=BLUE, bold=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        elif line.startswith('→'):
            # Summary paragraph
            cell.font = Font(size=9, name='Calibri', color='444444')
            cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.row_dimensions[row].height = 30
        elif line.startswith('⚠️'):
            # Warnings header
            cell.font = Font(bold=True, size=10, name='Calibri', color=RED)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        elif line.startswith('Per-account:') or line.startswith('Action items:'):
            # Section header
            cell.font = Font(bold=True, size=10, name='Calibri', color=NAV)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        elif line.startswith('  '):
            # Indented content (warnings, per-account, actions)
            cell.font = Font(size=8, name='Consolas', color='444444')  # Monospace for alignment
            cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.row_dimensions[row].height = 15
        else:
            # Generic line
            cell.font = Font(size=9, name='Calibri', color='444444')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

        row += 1

    row += 2  # Gap before table

    # ── Per-account table ──
    table_start_row = row

    # Headers
    headers = ['Account', 'Daily Spend', 'Monthly Spend', 'Daily Revenue', 'Monthly Revenue', 'ROAS', 'Inst. mROAS', 'Disc. mROAS vs Prev', 'Move']
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border()

    ws.row_dimensions[row].height = 28
    row += 1
    data_start_row = row

    # Data rows
    for i, alloc in enumerate(sorted(scenario.allocations, key=lambda a: a.account)):
        fill = PatternFill('solid', fgColor=LBLU if i % 2 == 0 else WHIT)

        cell = ws.cell(row=row, column=1, value=alloc.account)
        cell.fill = fill
        cell.border = _border()

        cell = ws.cell(row=row, column=2, value=alloc.daily_spend)
        cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()

        cell = ws.cell(row=row, column=3, value=alloc.monthly_spend)
        cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()

        cell = ws.cell(row=row, column=4, value=alloc.daily_revenue)
        cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()

        cell = ws.cell(row=row, column=5, value=alloc.monthly_revenue)
        cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()

        cell = ws.cell(row=row, column=6, value=alloc.roas)
        cell.number_format = '0.00"x"'
        cell.fill = fill
        cell.border = _border()

        cell = ws.cell(row=row, column=7, value=alloc.inst_mroas)
        cell.number_format = '0.00"x"'
        cell.fill = fill
        cell.border = _border()

        # Discrete mROAS vs prev - with clarification for D sheet when account was already at floor
        if alloc.discrete_mroas_vs_prev is not None:
            disc_val = alloc.discrete_mroas_vs_prev
            cell = ws.cell(row=row, column=8, value=disc_val)
            cell.number_format = '0.00"x"'
        else:
            # Check if this is Scenario D and account is at floor (inst_mroas ~= min_mroas)
            if scenario.id == "D" and abs(alloc.inst_mroas - min_mroas) < 0.1:
                disc_val = '— (already at floor)'
            else:
                disc_val = '—'
            cell = ws.cell(row=row, column=8, value=disc_val)
        cell.fill = fill
        cell.border = _border()

        cell = ws.cell(row=row, column=9, value=alloc.change_label if alloc.change_label else '—')
        cell.fill = fill
        cell.border = _border()
        # Color code move column
        if alloc.change_label and '▲' in alloc.change_label:
            cell.font = Font(color=GRN, bold=True, name='Calibri')
        elif alloc.change_label and '▼' in alloc.change_label:
            cell.font = Font(color=RED, bold=True, name='Calibri')

        row += 1

    data_end_row = row - 1

    # ── Apply conditional formatting to inst. mROAS and discrete mROAS columns ──
    # Inst. mROAS: column G
    ws.conditional_formatting.add(
        f'G{data_start_row}:G{data_end_row}',
        CellIsRule(operator='lessThan', formula=[str(min_mroas)], fill=PatternFill(start_color=RED, end_color=RED, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G{data_start_row}:G{data_end_row}',
        CellIsRule(operator='between', formula=[str(min_mroas), str(min_mroas + 1.5)], fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'G{data_start_row}:G{data_end_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=[str(min_mroas + 1.5)], fill=PatternFill(start_color=GRN, end_color=GRN, fill_type='solid'))
    )

    # Discrete mROAS: column H (same thresholds)
    ws.conditional_formatting.add(
        f'H{data_start_row}:H{data_end_row}',
        CellIsRule(operator='lessThan', formula=[str(min_mroas)], fill=PatternFill(start_color=RED, end_color=RED, fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H{data_start_row}:H{data_end_row}',
        CellIsRule(operator='between', formula=[str(min_mroas), str(min_mroas + 1.5)], fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'H{data_start_row}:H{data_end_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=[str(min_mroas + 1.5)], fill=PatternFill(start_color=GRN, end_color=GRN, fill_type='solid'))
    )

    # Set column widths
    ws.column_dimensions['A'].width = 24  # Account
    ws.column_dimensions['B'].width = 14  # Daily Spend
    ws.column_dimensions['C'].width = 16  # Monthly Spend
    ws.column_dimensions['D'].width = 14  # Daily Revenue
    ws.column_dimensions['E'].width = 16  # Monthly Revenue
    ws.column_dimensions['F'].width = 10  # ROAS
    ws.column_dimensions['G'].width = 12  # Inst. mROAS
    ws.column_dimensions['H'].width = 18  # Disc. mROAS vs Prev
    ws.column_dimensions['I'].width = 18  # Move
