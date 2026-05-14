"""
Diagnostic Excel sheet builders: Model Accuracy and CPC Diagnostics.
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

from budget_solver.constants import NAV, GRN, RED, LBLU, WHIT
from budget_solver.excel.styling import _hdr, _fmt_col, _border

# Amber not in constants — define locally
AMBR = 'FFC000'

# Error thresholds for traffic-light colouring
_ERR_GREEN = 0.10   # abs(error) < 10% → green
_ERR_AMBER = 0.20   # abs(error) < 20% → amber; ≥ 20% → red


def _err_fill(abs_err_pct: float) -> PatternFill:
    """Return green / amber / red fill based on absolute error magnitude."""
    if abs_err_pct < _ERR_GREEN:
        color = GRN
    elif abs_err_pct < _ERR_AMBER:
        color = AMBR
    else:
        color = RED
    return PatternFill('solid', fgColor=color)


def _data_font(bold=False):
    return Font(name='Calibri', size=9, bold=bold)


def _build_model_accuracy(wb, accuracy_df):
    """
    Build 'Model Accuracy' sheet from scored prediction history DataFrame.

    Expected columns: run_date, forecast_period, account_name,
    recommended_spend, predicted_revenue, actual_spend, actual_revenue,
    spend_error_pct, revenue_error_pct, revenue_abs_error_pct, actual_roas
    """
    ws = wb.create_sheet('Model Accuracy')
    ws.sheet_view.showGridLines = False

    if accuracy_df is None or (hasattr(accuracy_df, 'empty') and accuracy_df.empty):
        ws['A1'] = 'No model accuracy data available.'
        ws['A1'].font = Font(italic=True, size=9, name='Calibri', color='888888')
        return

    df = accuracy_df.copy()
    border = _border()
    row = 1

    # ── Title ──
    ws.cell(row=row, column=1, value='MODEL ACCURACY — HISTORICAL PREDICTION VS ACTUALS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, name='Calibri', color=NAV)
    row += 2

    # ── Summary block ──
    mape = df['revenue_abs_error_pct'].mean()
    bias = df['revenue_error_pct'].mean()
    n_months = df['forecast_period'].nunique()
    n_accounts = df['account_name'].nunique()

    summary_rows = [
        ('Portfolio MAPE (revenue)', f'{mape:.1%}'),
        ('Portfolio Bias (positive = under-predicted)', f'{bias:+.1%}'),
        ('Months scored', str(n_months)),
        ('Accounts scored', str(n_accounts)),
    ]
    for label, value in summary_rows:
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font = Font(bold=True, size=9, name='Calibri', color='444444')
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = _data_font()
        row += 1

    row += 1  # gap

    # ── Per-account summary table ──
    acct_hdrs = [
        'Account', 'Periods Scored', 'Mean Error %', 'Median Error %',
        'MAPE', 'Bias', 'Last Period', 'Last Error %',
    ]
    for c, h in enumerate(acct_hdrs, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    row += 1

    # Build per-account stats
    acct_stats = (
        df.groupby('account_name')
        .agg(
            periods_scored=('forecast_period', 'nunique'),
            mean_err=('revenue_error_pct', 'mean'),
            median_err=('revenue_error_pct', 'median'),
            mape=('revenue_abs_error_pct', 'mean'),
            bias=('revenue_error_pct', 'mean'),
        )
        .reset_index()
    )
    # Last period + last error per account
    last_rows = (
        df.sort_values('forecast_period')
        .groupby('account_name')
        .last()
        .reset_index()[['account_name', 'forecast_period', 'revenue_error_pct']]
        .rename(columns={'forecast_period': 'last_period', 'revenue_error_pct': 'last_err'})
    )
    acct_stats = acct_stats.merge(last_rows, on='account_name', how='left')
    acct_stats = acct_stats.sort_values('mape', ascending=False)

    for i, r in enumerate(acct_stats.itertuples(index=False)):
        fill = PatternFill('solid', fgColor=LBLU) if i % 2 == 1 else None
        row_data = [
            r.account_name,
            r.periods_scored,
            r.mean_err,
            r.median_err,
            r.mape,
            r.bias,
            r.last_period,
            r.last_err,
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = _data_font()
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')
            if fill:
                cell.fill = fill
            # Color-code error columns (3=mean_err, 4=median_err, 5=mape, 6=bias, 8=last_err)
            if c in (3, 4, 5, 6, 8) and val is not None:
                cell.fill = _err_fill(abs(float(val)))
            if c in (3, 4, 5, 6, 8):
                cell.number_format = '0.0%'
        row += 1

    row += 1  # gap

    # ── Detailed history table ──
    detail_hdrs = [
        'Forecast Period', 'Run Date', 'Account',
        'Predicted Spend', 'Actual Spend', 'Spend Error %',
        'Predicted Revenue', 'Actual Revenue', 'Revenue Error %', 'Actual ROAS',
    ]
    for c, h in enumerate(detail_hdrs, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    row += 1

    detail = df.sort_values(['forecast_period', 'account_name'], ascending=[False, True])

    for i, r in enumerate(detail.itertuples(index=False)):
        fill = PatternFill('solid', fgColor=LBLU) if i % 2 == 1 else None
        row_vals = [
            getattr(r, 'forecast_period', ''),
            getattr(r, 'run_date', ''),
            getattr(r, 'account_name', ''),
            getattr(r, 'recommended_spend', 0),
            getattr(r, 'actual_spend', 0),
            getattr(r, 'spend_error_pct', 0),
            getattr(r, 'predicted_revenue', 0),
            getattr(r, 'actual_revenue', 0),
            getattr(r, 'revenue_error_pct', 0),
            getattr(r, 'actual_roas', 0),
        ]
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = _data_font()
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if c > 3 else 'left')
            if fill:
                cell.fill = fill
            if c in (4, 5, 7, 8):
                cell.number_format = '€#,##0'
            if c in (6, 9):
                cell.number_format = '0.0%'
            if c == 10:
                cell.number_format = '0.00'
            # Traffic-light on Revenue Error %
            if c == 9 and val is not None:
                cell.fill = _err_fill(abs(float(val)))
        row += 1

    # Column widths
    widths = [14, 12, 24, 16, 16, 13, 18, 16, 15, 12]
    for c, w in enumerate(widths, 1):
        _fmt_col(ws, c, w)


def _build_cpc_diagnostics(wb, df, scenario_set):
    """
    Build 'CPC Diagnostics' sheet to visualise auction price dynamics.

    Args:
        wb: Workbook
        df: Raw daily input DataFrame with columns: account_name, date, cost, clicks, cpc
        scenario_set: ScenarioSet; scenarios[0]=A (current), scenarios[2]=C (recommended)
    """
    ws = wb.create_sheet('CPC Diagnostics')
    ws.sheet_view.showGridLines = False

    if df is None or df.empty or 'cpc' not in df.columns:
        ws['A1'] = 'No CPC data available.'
        ws['A1'].font = Font(italic=True, size=9, name='Calibri', color='888888')
        return

    border = _border()
    row = 1

    # ── Title ──
    ws.cell(row=row, column=1, value='CPC DIAGNOSTICS — AUCTION PRICE DYNAMICS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, name='Calibri', color=NAV)
    row += 1

    subtitle = (
        'CPC tends to rise with spend as Smart Bidding pushes into more competitive auctions. '
        'High CPC inflation = model revenue curve may be over-optimistic at higher spend levels. '
        'Use for qualitative interpretation; CPC is not directly used in the optimizer.'
    )
    cell = ws.cell(row=row, column=1, value=subtitle)
    cell.font = Font(italic=True, size=9, name='Calibri', color='888888')
    cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 2

    # ── Prepare data ──
    work = df.copy()
    work['date'] = pd.to_datetime(work['date'])
    max_date = work['date'].max()
    cutoff = max_date - pd.Timedelta(days=30)

    last30 = work[work['date'] > cutoff]
    full = work

    # Scenario allocations
    scen_a = scenario_set.scenarios[0]
    scen_c = scenario_set.scenarios[2]

    def _alloc(scenario, account):
        for a in getattr(scenario, 'allocations', []):
            if a.account == account:
                return a.monthly_spend
        return 0

    accounts = sorted(work['account_name'].unique())

    # ── Per-account summary table ──
    sum_hdrs = [
        'Account', 'Avg CPC Last 30d', 'Avg CPC Full Period',
        'CPC Trend %', 'Spend Change %', 'Recommended Spend', 'Current Spend',
    ]
    for c, h in enumerate(sum_hdrs, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    row += 1

    for i, acc in enumerate(accounts):
        acc_full = full[full['account_name'] == acc]['cpc']
        acc_last = last30[last30['account_name'] == acc]['cpc']

        cpc_full = acc_full.mean() if len(acc_full) else 0.0
        cpc_last = acc_last.mean() if len(acc_last) else 0.0
        cpc_trend = (cpc_last / cpc_full - 1) if cpc_full else 0.0

        curr_spend = _alloc(scen_a, acc)
        rec_spend = _alloc(scen_c, acc)
        spend_chg = (rec_spend / curr_spend - 1) if curr_spend else 0.0

        # Amber flag: spend increasing AND CPC inflation > +15%
        cpc_inflation_risk = spend_chg > 0 and cpc_trend > 0.15

        row_vals = [acc, cpc_last, cpc_full, cpc_trend, spend_chg, rec_spend, curr_spend]
        alt_fill = PatternFill('solid', fgColor=LBLU) if i % 2 == 1 else None

        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = _data_font()
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')
            if alt_fill:
                cell.fill = alt_fill
            if c in (2, 3):
                cell.number_format = '€0.00'
            if c in (4, 5):
                cell.number_format = '+0.0%;-0.0%;0.0%'
            if c in (6, 7):
                cell.number_format = '€#,##0'
            # Amber highlight on CPC Trend column when inflation risk detected
            if c == 4 and cpc_inflation_risk:
                cell.fill = PatternFill('solid', fgColor=AMBR)
        row += 1

    row += 1  # gap

    # ── Monthly CPC trend table (wide format, last 6 months) ──
    ws.cell(row=row, column=1, value='MONTHLY CPC TREND (last 6 months)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=10, name='Calibri', color=NAV)
    row += 1

    work['month'] = work['date'].dt.to_period('M').astype(str)
    recent_months = sorted(work['month'].unique())[-6:]

    # Header row
    ws.cell(row=row, column=1, value='Account')
    ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=NAV)
    ws.cell(row=row, column=1).font = Font(bold=True, size=9, name='Calibri', color=WHIT)
    ws.cell(row=row, column=1).border = border

    for c, month in enumerate(recent_months, 2):
        cell = ws.cell(row=row, column=c, value=month)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1

    for i, acc in enumerate(accounts):
        alt_fill = PatternFill('solid', fgColor=LBLU) if i % 2 == 1 else None
        cell = ws.cell(row=row, column=1, value=acc)
        cell.font = _data_font()
        cell.border = border
        if alt_fill:
            cell.fill = alt_fill

        for c, month in enumerate(recent_months, 2):
            subset = work[(work['account_name'] == acc) & (work['month'] == month)]['cpc']
            median_cpc = subset.median() if len(subset) else None
            cell = ws.cell(row=row, column=c, value=median_cpc)
            cell.font = _data_font()
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if median_cpc is not None:
                cell.number_format = '€0.00'
            if alt_fill:
                cell.fill = alt_fill
        row += 1

    # Column widths — summary table cols first, then monthly table cols
    summary_widths = [24, 16, 17, 13, 14, 18, 16]
    for c, w in enumerate(summary_widths, 1):
        _fmt_col(ws, c, w)
    # Monthly trend table: col 1 already set above; set month columns
    for c in range(2, len(recent_months) + 2):
        _fmt_col(ws, c, 13)
