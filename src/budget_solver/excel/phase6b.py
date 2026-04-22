"""
Phase 6b supplementary sheets: Extended Budget, Curve Diagnostics, Outlier Log, Demand Index.
"""
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference, marker

from budget_solver.constants import NAV, BLUE, GRN, RED, LBLU, WHIT, WEEKS_PER_MONTH
from budget_solver.excel.styling import _hdr, _fmt_col, _border
from budget_solver.solver import optimize_budget


def _build_extended_budget(wb, scenario_set, steps=6):
    """
    Build Extended Budget sheet — efficient frontier sweep from C → D.

    Sweep from C's budget to D's budget in equal increments, re-running solver at each step.
    First row = C (already computed), last row = D (already computed).
    Middle rows = new solver runs.

    Args:
        wb: Workbook
        scenario_set: ScenarioSet
        steps: Number of rows (default 6: C, C+1/5, ..., D)
    """
    ws = wb.create_sheet('Extended Budget')
    ws.sheet_view.showGridLines = False

    scenarios = scenario_set.scenarios
    scen_c = scenarios[2]
    scen_d = scenarios[3] if len(scenarios) > 3 else None

    if not scen_d:
        # No D scenario, can't sweep
        ws['A1'] = 'Extended Budget sheet requires Scenario D (no D generated)'
        ws['A1'].font = Font(italic=True, size=9, name='Calibri', color='888888')
        return

    predict_fns = scenario_set.predict_fns
    min_mroas = scenario_set.min_mroas
    accounts = sorted(predict_fns.keys())

    row = 1

    # ── Title ──
    ws.cell(row=row, column=1, value='EXTENDED BUDGET — EFFICIENT FRONTIER')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, name='Calibri', color=NAV)
    row += 1

    ws.cell(row=row, column=1, value='Sweep from C (recommended) → D (max justified) in equal increments. Shows diminishing returns as budget scales.')
    ws.cell(row=row, column=1).font = Font(size=9, name='Calibri', color='888888')
    row += 2

    # ── Table headers ──
    headers = ['Budget', 'Revenue', 'Blended ROAS', 'Incremental Budget', 'Incremental Revenue', 'Incremental ROAS']
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border()

    row += 1
    data_start_row = row

    # ── Generate sweep ──
    budget_c = scen_c.budget_monthly
    budget_d = scen_d.budget_monthly
    gap = budget_d - budget_c

    sweep_budgets = [budget_c + (i / (steps - 1)) * gap for i in range(steps)]

    # Collect sweep data
    sweep_results = []

    # Build bounds for solver (all accounts at breakeven floor)
    min_spend = {acc: 0 for acc in accounts}
    max_spend = {}
    for acc in accounts:
        fn, params, _, _ = scenario_set.model_info[acc]
        a = params[0]
        weekly_breakeven = a / min_mroas
        monthly_breakeven = weekly_breakeven * WEEKS_PER_MONTH
        max_spend[acc] = monthly_breakeven

    for i, budget in enumerate(sweep_budgets):
        if i == 0:
            # First row = C's allocation (already computed)
            alloc = {a.account: a.monthly_spend for a in scen_c.allocations}
            revenue = scen_c.revenue_monthly
        elif i == steps - 1:
            # Last row = D's allocation (already computed)
            alloc = {a.account: a.monthly_spend for a in scen_d.allocations}
            revenue = scen_d.revenue_monthly
        else:
            # Middle rows: run actual solver at this budget level
            # This gives the true efficient frontier (concave curve showing diminishing returns)
            alloc, _, revenue = optimize_budget(
                predict_fns=predict_fns,
                total_budget=budget,
                min_spend=min_spend,
                max_spend=max_spend
            )
            # Revenue already computed by optimize_budget as return value

        blended_roas = revenue / budget if budget > 0 else 0

        sweep_results.append({
            'budget': budget,
            'revenue': revenue,
            'roas': blended_roas,
            'alloc': alloc
        })

    # Calculate incremental metrics
    for i, result in enumerate(sweep_results):
        if i == 0:
            result['inc_budget'] = None
            result['inc_revenue'] = None
            result['inc_roas'] = None
        else:
            prev = sweep_results[i - 1]
            result['inc_budget'] = result['budget'] - prev['budget']
            result['inc_revenue'] = result['revenue'] - prev['revenue']
            result['inc_roas'] = result['inc_revenue'] / result['inc_budget'] if result['inc_budget'] > 0 else 0

    # ── Write data rows ──
    for i, result in enumerate(sweep_results):
        fill = PatternFill('solid', fgColor=LBLU if i % 2 == 0 else WHIT)

        # Mark C and D rows
        is_c = (i == 0)
        is_d = (i == steps - 1)
        row_font = Font(bold=True, name='Calibri') if (is_c or is_d) else Font(name='Calibri')

        cell = ws.cell(row=row, column=1, value=result['budget'])
        cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()
        cell.font = row_font

        cell = ws.cell(row=row, column=2, value=result['revenue'])
        cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()
        cell.font = row_font

        cell = ws.cell(row=row, column=3, value=result['roas'])
        cell.number_format = '0.00"x"'
        cell.fill = fill
        cell.border = _border()
        cell.font = row_font

        cell = ws.cell(row=row, column=4, value=result['inc_budget'] if result['inc_budget'] is not None else '—')
        if result['inc_budget'] is not None:
            cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()
        cell.font = row_font

        cell = ws.cell(row=row, column=5, value=result['inc_revenue'] if result['inc_revenue'] is not None else '—')
        if result['inc_revenue'] is not None:
            cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()
        cell.font = row_font

        cell = ws.cell(row=row, column=6, value=result['inc_roas'] if result['inc_roas'] is not None else '—')
        if result['inc_roas'] is not None:
            cell.number_format = '0.00"x"'
        cell.fill = fill
        cell.border = _border()
        cell.font = row_font

        row += 1

    data_end_row = row - 1

    # ── Chart: Budget vs Revenue (primary) and Incremental ROAS (secondary) ──
    chart = LineChart()
    chart.title = 'Efficient Frontier: Budget → Revenue + Incremental ROAS'
    chart.style = 10
    chart.y_axis.title = 'Revenue (€)'
    chart.y_axis.numFmt = '€#,##0'
    chart.x_axis.title = 'Budget (€)'
    chart.x_axis.numFmt = '€#,##0'
    chart.height = 14
    chart.width = 22

    # Revenue line (primary y-axis)
    rev_ref = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
    budget_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
    chart.add_data(rev_ref, titles_from_data=True)
    chart.set_categories(budget_ref)
    chart.series[0].graphicalProperties.line.solidFill = BLUE
    chart.series[0].graphicalProperties.line.width = 25000

    # Incremental ROAS line (secondary y-axis)
    inc_roas_chart = LineChart()
    inc_roas_ref = Reference(ws, min_col=6, min_row=data_start_row - 1, max_row=data_end_row)
    inc_roas_chart.add_data(inc_roas_ref, titles_from_data=True)
    inc_roas_chart.set_categories(budget_ref)
    inc_roas_chart.y_axis.axId = 200
    inc_roas_chart.y_axis.title = 'Incremental ROAS'
    inc_roas_chart.y_axis.crosses = 'max'
    inc_roas_chart.y_axis.numFmt = '0.00"x"'
    inc_roas_chart.series[0].graphicalProperties.line.solidFill = 'ED7D31'  # Orange
    inc_roas_chart.series[0].graphicalProperties.line.width = 20000
    inc_roas_chart.series[0].graphicalProperties.line.dashStyle = 'dash'

    chart += inc_roas_chart

    ws.add_chart(chart, f'H{data_start_row - 2}')

    # Set column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16


def _build_curve_diagnostics(wb, scenario_set, model_info, account_data):
    """
    Build Curve Diagnostics sheet — per-account R², training window, breakeven, mini charts.

    Shows fitted curve from 0 to D's spend, with vertical markers at A/B/C/D.
    """
    ws = wb.create_sheet('Curve Diagnostics')
    ws.sheet_view.showGridLines = False

    scenarios = scenario_set.scenarios
    scen_a = scenarios[0]
    scen_b = scenarios[1]
    scen_c = scenarios[2]
    scen_d = scenarios[3] if len(scenarios) > 3 else None

    predict_fns = scenario_set.predict_fns
    min_mroas = scenario_set.min_mroas
    accounts = sorted(predict_fns.keys())

    row = 1

    # ── Title ──
    ws.cell(row=row, column=1, value='CURVE DIAGNOSTICS — MODEL FIT & BREAKEVEN ANALYSIS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, name='Calibri', color=NAV)
    row += 2

    # ── Table headers ──
    headers = ['Account', 'Model', 'R²', 'Data Points', 'Calibration Factor', 'Training Window', f'Breakeven @ {min_mroas:.1f}x']
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border()

    row += 1

    # ── Data rows + mini charts ──
    from budget_solver.constants import WEEKS_PER_MONTH

    for i, account in enumerate(accounts):
        chart_anchor_row = row

        fn, params, r2, mname = model_info[account]
        a = params[0]

        # Breakeven at min_mroas
        weekly_breakeven = a / min_mroas
        monthly_breakeven = weekly_breakeven * WEEKS_PER_MONTH

        # Training data
        data = account_data[account]
        n_points = len(data['spend'])
        avg_spend = float(np.mean(data['spend']))

        # Calibration factor (assumed 1.0 if not stored - would need to extract from scenario generation)
        cal_factor = 1.0  # TODO: extract from scenario_set if available

        # Training window (assumed 6 months if not stored)
        training_window = '6 months'  # TODO: extract from scenario_set if available

        fill = PatternFill('solid', fgColor=LBLU if i % 2 == 0 else WHIT)

        ws.cell(row=row, column=1, value=account).fill = fill
        ws.cell(row=row, column=1).border = _border()

        ws.cell(row=row, column=2, value=mname).fill = fill
        ws.cell(row=row, column=2).border = _border()

        cell = ws.cell(row=row, column=3, value=r2)
        cell.number_format = '0.00'
        cell.fill = fill
        cell.border = _border()

        ws.cell(row=row, column=4, value=n_points).fill = fill
        ws.cell(row=row, column=4).border = _border()

        cell = ws.cell(row=row, column=5, value=cal_factor)
        cell.number_format = '0.00"x"'
        cell.fill = fill
        cell.border = _border()

        ws.cell(row=row, column=6, value=training_window).fill = fill
        ws.cell(row=row, column=6).border = _border()

        cell = ws.cell(row=row, column=7, value=monthly_breakeven)
        cell.number_format = '€#,##0'
        cell.fill = fill
        cell.border = _border()

        # ── Mini chart for this account ──
        # Curve from 0 to D's spend (or 2x current max if no D)
        alloc_a = next((a for a in scen_a.allocations if a.account == account), None)
        alloc_b = next((a for a in scen_b.allocations if a.account == account), None)
        alloc_c = next((a for a in scen_c.allocations if a.account == account), None)
        alloc_d = next((a for a in scen_d.allocations if a.account == account), None) if scen_d else None

        max_spend = alloc_d.monthly_spend if alloc_d else (alloc_c.monthly_spend * 2)

        # Generate curve points
        curve_points = np.linspace(0, max_spend, 20)
        curve_start_col = 9  # Column I
        curve_data_row = row

        # Write curve data in hidden columns
        for j, spend_val in enumerate(curve_points):
            col = curve_start_col + j
            revenue_val = predict_fns[account](spend_val)
            ws.cell(row=curve_data_row, column=col, value=revenue_val)

        # TODO: Add mini line chart here (simplified for now - charts are complex in openpyxl)
        # For now, just show the table data. Full chart implementation would require:
        # - LineChart with data from columns I onwards
        # - Vertical line markers at A/B/C/D positions (requires manual series or scatter overlay)

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16


def _build_outlier_log(wb, removal_log):
    """Build Outlier Log sheet (ported from old implementation)."""
    ws = wb.create_sheet('Outlier Log')
    ws.sheet_view.showGridLines = False

    row = 1

    ws.cell(row=row, column=1, value='OUTLIER & ANOMALY REMOVAL LOG')
    ws.cell(row=row, column=1).font = Font(bold=True, size=13, name='Calibri', color=NAV)
    row += 1

    ws.cell(row=row, column=1, value='Weeks excluded before fitting response curves. Review these to confirm removals are appropriate.')
    ws.cell(row=row, column=1).font = Font(size=9, name='Calibri', color='888888')
    row += 2

    log_cols = ['Account', 'Week', 'Spend (€)', 'Revenue (€)', 'ROAS', 'Reason']
    for c, h in enumerate(log_cols, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, size=9, name='Calibri', color=WHIT)
        cell.fill = PatternFill('solid', fgColor=NAV)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border()

    row += 1

    if removal_log:
        for i, entry in enumerate(removal_log):
            fill = PatternFill('solid', fgColor=LBLU if i % 2 == 0 else WHIT)
            vals = [entry['account'], entry['week'], entry['spend'],
                    entry['revenue'], entry['roas'], entry['reason']]
            fmts = ['@', '@', '€#,##0', '€#,##0', '0.00', '@']
            for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.number_format = fmt
                cell.fill = fill
                cell.border = _border()
            row += 1
    else:
        ws.cell(row=row, column=1, value='No outliers removed.')
        ws.cell(row=row, column=1).font = Font(italic=True, color='888888', name='Calibri')

    log_widths = [28, 24, 14, 14, 10, 55]
    for c, w in enumerate(log_widths, 1):
        _fmt_col(ws, c, w)


def _build_demand_index(wb, demand_index, demand_normalized, forecast_week, forecast_label):
    """Build Demand Index sheet (ported from old implementation)."""
    ws = wb.create_sheet('Demand Index')
    ws.sheet_view.showGridLines = False

    row = 1

    ws.cell(row=row, column=1, value='SEASONAL DEMAND INDEX')
    ws.cell(row=row, column=1).font = Font(bold=True, size=13, name='Calibri', color=NAV)
    row += 1

    norm_note = 'demand-normalized before fitting' if demand_normalized else 'diagnostic only — curves were NOT normalized'
    ws.cell(row=row, column=1, value=(
        f'Weekly demand multiplier (mean = 1.0). {norm_note}. '
        f'Forecast period: {forecast_label} (week {forecast_week}).'
    ))
    ws.cell(row=row, column=1).font = Font(size=9, name='Calibri', color='888888')
    row += 2

    ws.cell(row=row, column=1, value='ISO Week')
    ws.cell(row=row, column=2, value='Demand Multiplier')
    ws.cell(row=row, column=3, value='Relative to avg')
    _hdr(ws, row, 3)

    row += 1

    if demand_index:
        for i, wk in enumerate(range(1, 54)):
            fill = PatternFill('solid', fgColor=LBLU if i % 2 == 0 else WHIT)
            mult = demand_index.get(wk, 1.0)
            diff = mult - 1.0

            cell = ws.cell(row=row, column=1, value=wk)
            cell.number_format = '0'
            cell.fill = fill
            cell.border = _border()

            cell = ws.cell(row=row, column=2, value=round(mult, 4))
            cell.number_format = '0.000'
            cell.fill = fill
            cell.border = _border()
            if wk == forecast_week:
                cell.font = Font(bold=True, color=BLUE, name='Calibri')

            cell = ws.cell(row=row, column=3, value=round(diff, 4))
            cell.number_format = '+0.0%;-0.0%;0.0%'
            cell.fill = fill
            cell.border = _border()
            cell.font = Font(color=GRN if diff >= 0 else RED, name='Calibri')

            row += 1

        # Bar chart
        from openpyxl.chart import BarChart
        chart = BarChart()
        chart.title = 'Seasonal Demand Index by Week'
        chart.style = 10
        chart.y_axis.title = 'Demand Multiplier'
        chart.x_axis.title = 'ISO Week'
        chart.height = 14
        chart.width = 24
        data_ref = Reference(ws, min_col=2, min_row=4, max_row=4 + 52)
        cats_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + 52)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, 'E4')

    for c, w in enumerate([12, 20, 18], 1):
        _fmt_col(ws, c, w)
